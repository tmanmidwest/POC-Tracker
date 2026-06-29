"""End-to-end UI tests for the POC Tracker web app."""

from __future__ import annotations

import struct
import zlib

import pytest
from fastapi.testclient import TestClient


def _login(client: TestClient, username: str, password: str) -> None:
    resp = client.post(
        "/ui/login",
        data={"username": username, "password": password},
        follow_redirects=False,
    )
    assert resp.status_code == 303, resp.text


@pytest.fixture
def ui(client: TestClient) -> TestClient:
    from app.config import get_settings

    s = get_settings()
    _login(client, s.initial_admin_username, s.initial_admin_password)
    return client


def _png() -> bytes:
    sig = b"\x89PNG\r\n\x1a\n"

    def chunk(t: bytes, d: bytes) -> bytes:
        c = t + d
        return struct.pack(">I", len(d)) + c + struct.pack(">I", zlib.crc32(c) & 0xFFFFFFFF)

    ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
    idat = zlib.compress(b"\x00\xff\x00\x00")
    return sig + chunk(b"IHDR", ihdr) + chunk(b"IDAT", idat) + chunk(b"IEND", b"")


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------


def test_unauthed_redirects_to_login(client: TestClient) -> None:
    resp = client.get("/ui/dashboard", follow_redirects=False)
    assert resp.status_code == 303
    assert "/ui/login" in resp.headers["location"]


def test_dashboard_renders(ui: TestClient) -> None:
    resp = ui.get("/ui/dashboard")
    assert resp.status_code == 200
    assert "Dashboard" in resp.text


def test_dashboard_scope_defaults_to_mine_and_is_sticky(ui: TestClient) -> None:
    """Internal users default to "My POCs" (projects they're the SE on) and the
    My/All choice sticks across visits."""
    from app.config import get_settings
    from app.db import get_session_factory
    from app.models import AppUser, Project

    db = get_session_factory()()
    me = db.query(AppUser).filter(
        AppUser.username == get_settings().initial_admin_username
    ).one()

    cid = _create_customer(ui, "Scope Co")
    mine_id = _create_project(ui, cid, "Mine POC")
    other_id = _create_project(ui, cid, "Other POC")
    db.get(Project, mine_id).sales_engineer_id = me.id  # assigned to me
    db.commit()

    # Default scope = mine: only my project shows.
    page = ui.get("/ui/dashboard").text
    assert "Mine POC" in page
    assert "Other POC" not in page

    # Switching to all shows both...
    page = ui.get("/ui/dashboard?scope=all").text
    assert "Mine POC" in page and "Other POC" in page

    # ...and the choice is sticky: a plain visit still shows all.
    assert "Other POC" in ui.get("/ui/dashboard").text

    # Flipping back to mine sticks too.
    assert "Other POC" not in ui.get("/ui/dashboard?scope=mine").text
    assert "Other POC" not in ui.get("/ui/dashboard").text


def test_dashboard_scope_unassigned_and_specific_engineer(ui: TestClient) -> None:
    """Scope supports "unassigned" and a specific sales engineer ("user:<id>")."""
    from app.db import get_session_factory
    from app.models import AppUser, Project
    from app.services.passwords import hash_password

    db = get_session_factory()()
    # A teammate who will own one project.
    teammate = AppUser(username="se_jane", password_hash=hash_password("password123"),
                       is_active=True, is_admin=False, display_name="Jane Doe")
    db.add(teammate)
    db.commit()

    cid = _create_customer(ui, "Assign Co")
    jane_id = _create_project(ui, cid, "Jane POC")
    free_id = _create_project(ui, cid, "Unowned POC")
    db.get(Project, jane_id).sales_engineer_id = teammate.id
    db.commit()

    # Filter to a specific engineer shows only their project.
    page = ui.get(f"/ui/dashboard?scope=user:{teammate.id}").text
    assert "Jane POC" in page
    assert "Unowned POC" not in page
    # The engineer appears as an option in the filter.
    assert "Jane Doe" in page

    # Filter to unassigned shows only the project with no sales engineer.
    page = ui.get("/ui/dashboard?scope=unassigned").text
    assert "Unowned POC" in page
    assert "Jane POC" not in page

    # A malformed scope value falls back safely (no 500).
    assert ui.get("/ui/dashboard?scope=user:notanumber").status_code == 200


def test_root_redirects_to_dashboard(ui: TestClient) -> None:
    resp = ui.get("/", follow_redirects=False)
    assert resp.headers["location"] == "/ui/dashboard"


# ---------------------------------------------------------------------------
# Customer + project + use-case flow
# ---------------------------------------------------------------------------


def _create_customer(ui: TestClient, name: str) -> int:
    resp = ui.post(
        "/ui/customers/new",
        data={"name": name, "website": "", "notes": ""},
        follow_redirects=False,
    )
    assert resp.status_code == 303
    return int(resp.headers["location"].rsplit("/", 1)[1])


def _create_project(ui: TestClient, customer_id: int, name: str) -> int:
    resp = ui.post(
        "/ui/projects/new",
        data={
            "customer_id": customer_id, "name": name, "status_id": "",
            "start_date": "", "end_date": "", "sales_engineer_id": "",
            "account_executive": "", "account_executive_email": "", "notes": "",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 303
    return int(resp.headers["location"].rsplit("/", 1)[1])


def test_full_project_flow(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import ProjectUseCase

    cid = _create_customer(ui, "Globex")
    pid = _create_project(ui, cid, "Globex POC")

    # Add two library entries, then re-open the picker and add the same two
    # plus one more — only the new one should be added (de-dup).
    ui.post(f"/ui/projects/{pid}/use-cases/from-library",
            data={"library_ids": ["1", "2"]}, follow_redirects=False)
    ui.post(f"/ui/projects/{pid}/use-cases/from-library",
            data={"library_ids": ["1", "2", "3"]}, follow_redirects=False)

    db = get_session_factory()()
    count = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).count()
    assert count == 3

    # Add a custom (ad-hoc) use case.
    ui.post(f"/ui/projects/{pid}/use-cases",
            data={"category": "Custom", "name": "Client special", "reference_number": "9.1",
                  "description": "", "success_validation": "", "feature_type_id": ""},
            follow_redirects=False)
    count = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).count()
    assert count == 4

    page = ui.get(f"/ui/projects/{pid}")
    assert page.status_code == 200
    assert "Client special" in page.text


def test_salesforce_opp_link(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import Project

    cid = _create_customer(ui, "SF Customer")
    # No-scheme URL should be normalized to https://, and render as a tidy link.
    ui.post("/ui/projects/new", data={
        "customer_id": cid, "name": "SF POC", "status_id": "",
        "start_date": "", "end_date": "", "sales_engineer_id": "",
        "account_executive": "", "account_executive_email": "",
        "salesforce_opp_url": "acme.lightning.force.com/opp/1",
        "notebook_url": "notebooks.example.com/n/42",
        "poc_instance_url": "poc.example.com/i/7", "notes": "",
    }, follow_redirects=False)
    db = get_session_factory()()
    p = db.query(Project).filter(Project.name == "SF POC").one()
    assert p.salesforce_opp_url == "https://acme.lightning.force.com/opp/1"
    assert p.notebook_url == "https://notebooks.example.com/n/42"
    assert p.poc_instance_url == "https://poc.example.com/i/7"

    detail = ui.get(f"/ui/projects/{p.id}").text
    assert "Salesforce Opp" in detail and p.salesforce_opp_url in detail
    assert "Notebook Link" in detail and p.notebook_url in detail
    assert "POC Instance" in detail and p.poc_instance_url in detail
    # The dashboard defaults to "My POCs"; this project has no assigned sales
    # engineer, so check the "All POCs" scope to see it.
    assert "Salesforce Opp" in ui.get("/ui/dashboard?scope=all").text

    # A javascript: (or any non-http) scheme is rejected, not stored.
    ui.post(f"/ui/projects/{p.id}/edit", data={
        "customer_id": cid, "name": "SF POC", "status_id": str(p.status_id),
        "start_date": "", "end_date": "", "sales_engineer_id": "",
        "account_executive": "", "account_executive_email": "",
        "salesforce_opp_url": "javascript:alert(1)", "notes": "",
    }, follow_redirects=False)
    db.expire_all()
    assert db.get(Project, p.id).salesforce_opp_url is None


def test_screenshot_upload_and_serve(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import ProjectUseCase, Screenshot

    cid = _create_customer(ui, "Initech")
    pid = _create_project(ui, cid, "Initech POC")
    ui.post(f"/ui/projects/{pid}/use-cases/from-library",
            data={"library_ids": ["1"]}, follow_redirects=False)
    db = get_session_factory()()
    uc = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).first()

    resp = ui.post(f"/ui/projects/use-cases/{uc.id}/screenshots",
                   files={"file": ("s.png", _png(), "image/png")},
                   data={"caption": "login"}, follow_redirects=False)
    assert resp.status_code == 303
    shot = db.query(Screenshot).first()
    assert shot is not None
    served = ui.get(f"/ui/projects/screenshots/{shot.id}")
    assert served.status_code == 200
    assert served.headers["content-type"] == "image/png"


def test_reject_non_image_screenshot(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import ProjectUseCase, Screenshot

    cid = _create_customer(ui, "Hooli")
    pid = _create_project(ui, cid, "Hooli POC")
    ui.post(f"/ui/projects/{pid}/use-cases/from-library",
            data={"library_ids": ["1"]}, follow_redirects=False)
    db = get_session_factory()()
    uc = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).first()
    ui.post(f"/ui/projects/use-cases/{uc.id}/screenshots",
            files={"file": ("notes.txt", b"hello", "text/plain")}, follow_redirects=False)
    assert db.query(Screenshot).count() == 0


def test_project_notes_crud_and_attachments(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import NoteAttachment, ProjectNote

    cid = _create_customer(ui, "Notes Co")
    pid = _create_project(ui, cid, "Notes POC")

    # Add a note with a date and a PDF attachment in one shot.
    resp = ui.post(
        f"/ui/projects/{pid}/notes",
        data={"body": "Kicked off the POC", "note_date": "2026-06-20"},
        files=[("files", ("plan.pdf", b"%PDF-1.4 fake", "application/pdf"))],
        follow_redirects=False,
    )
    assert resp.status_code == 303

    db = get_session_factory()()
    note = db.query(ProjectNote).filter(ProjectNote.project_id == pid).one()
    assert note.body == "Kicked off the POC"
    assert note.note_date.isoformat() == "2026-06-20"
    assert note.created_by  # stamped with the acting user
    att = db.query(NoteAttachment).filter(NoteAttachment.project_note_id == note.id).one()
    assert att.original_filename == "plan.pdf"
    att_id = att.id

    # The note + attachment link render on the project page.
    page = ui.get(f"/ui/projects/{pid}").text
    assert "Kicked off the POC" in page
    assert f"/ui/projects/note-attachments/{att_id}" in page

    # Attachment is served (inline) for viewing/download.
    served = ui.get(f"/ui/projects/note-attachments/{att_id}")
    assert served.status_code == 200
    assert "inline" in served.headers.get("content-disposition", "")

    # Edit the note's body + date.
    ui.post(f"/ui/projects/notes/{note.id}/edit",
            data={"body": "Updated summary", "note_date": "2026-06-21"},
            follow_redirects=False)
    db.expire_all()
    note = db.get(ProjectNote, note.id)
    assert note.body == "Updated summary"
    assert note.note_date.isoformat() == "2026-06-21"

    # Upload a second attachment to the existing note.
    ui.post(f"/ui/projects/notes/{note.id}/attachments",
            files=[("files", ("doc.docx", b"PK\x03\x04 fake", None))],
            follow_redirects=False)
    assert db.query(NoteAttachment).filter(NoteAttachment.project_note_id == note.id).count() == 2

    # A disallowed file type is rejected (count unchanged).
    ui.post(f"/ui/projects/notes/{note.id}/attachments",
            files=[("files", ("evil.exe", b"MZ", "application/octet-stream"))],
            follow_redirects=False)
    assert db.query(NoteAttachment).filter(NoteAttachment.project_note_id == note.id).count() == 2

    # Remove one attachment.
    ui.post(f"/ui/projects/note-attachments/{att_id}/delete", follow_redirects=False)
    assert db.query(NoteAttachment).filter(NoteAttachment.id == att_id).count() == 0

    # Deleting the note cascades to its remaining attachments.
    ui.post(f"/ui/projects/notes/{note.id}/delete", follow_redirects=False)
    assert db.query(ProjectNote).filter(ProjectNote.project_id == pid).count() == 0
    assert db.query(NoteAttachment).filter(NoteAttachment.project_note_id == note.id).count() == 0


def test_project_report_html_and_zip(ui: TestClient) -> None:
    import io
    import zipfile

    from app.db import get_session_factory
    from app.models import ProjectUseCase

    cid = _create_customer(ui, "Report Co")
    pid = _create_project(ui, cid, "Report POC")
    ui.post(f"/ui/projects/{pid}/use-cases/from-library",
            data={"library_ids": ["1"]}, follow_redirects=False)
    db = get_session_factory()()
    uc = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).first()
    ui.post(f"/ui/projects/use-cases/{uc.id}/screenshots",
            files={"file": ("shot.png", _png(), "image/png")}, follow_redirects=False)
    ui.post(f"/ui/projects/{pid}/notes",
            data={"body": "Journal entry one", "note_date": "2026-06-20"},
            files=[("files", ("brief.pdf", b"%PDF-1.4 x", "application/pdf"))],
            follow_redirects=False)

    # Standalone report page: no nav sidebar, shows journal + screenshot + zip button.
    r = ui.get(f"/ui/reports/projects/{pid}")
    assert r.status_code == 200
    assert "sidebar__nav" not in r.text  # navigation bar omitted
    assert "Journal entry one" in r.text
    assert "/ui/projects/screenshots/" in r.text
    assert "Download all (.zip)" in r.text  # has_artifacts -> zip button shown
    assert "artifacts.zip?v=" in r.text  # cache-busted download link
    assert "/pdf?v=" in r.text

    # Zip bundles screenshots + attachments (and the PDF when WeasyPrint is present).
    z = ui.get(f"/ui/reports/projects/{pid}/artifacts.zip")
    assert z.status_code == 200
    assert z.headers["content-type"] == "application/zip"
    assert "no-store" in z.headers.get("cache-control", "")  # never browser-cached
    names = zipfile.ZipFile(io.BytesIO(z.content)).namelist()
    assert any("/screenshots/" in n for n in names), names
    assert any("/attachments/" in n for n in names), names


def test_project_report_pdf(ui: TestClient) -> None:
    # Needs system libs (pango/glib); runs in the container/CI. Importing
    # WeasyPrint raises ImportError when the Python package is missing and
    # OSError when its native libraries can't be loaded — skip on either.
    try:
        import weasyprint  # noqa: F401
    except (ImportError, OSError):
        pytest.skip("WeasyPrint or its native libraries are unavailable")

    cid = _create_customer(ui, "PDF Co")
    pid = _create_project(ui, cid, "PDF POC")
    r = ui.get(f"/ui/reports/projects/{pid}/pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:5] == b"%PDF-"


def test_use_case_view_prefs(ui: TestClient) -> None:
    import json

    from app.db import get_session_factory
    from app.models import ProjectUseCase, UseCaseStatus, UseCaseViewPref

    cid = _create_customer(ui, "View Co")
    pid = _create_project(ui, cid, "View POC")
    ui.post(f"/ui/projects/{pid}/use-cases/from-library",
            data={"library_ids": ["1", "2"]}, follow_redirects=False)
    db = get_session_factory()()
    ucs = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).all()
    complete = db.query(UseCaseStatus).filter(UseCaseStatus.is_complete_status.is_(True)).first()
    ucs[1].success_validation = "Must pass SSO test"
    db.commit()
    # Mark the first use case complete (leaves one open).
    ui.post(f"/ui/projects/use-cases/{ucs[0].id}/status",
            data={"status_id": complete.id}, follow_redirects=False)

    # Default view: the Success validation field is off, so its value appears
    # only in the (always-present) edit modal textarea — exactly once.
    page = ui.get(f"/ui/projects/{pid}").text
    assert page.count("Must pass SSO test") == 1

    # Enable the Success validation field and filter to not-completed.
    ui.post(f"/ui/projects/{pid}/use-case-view",
            data={"field_success_validation": "1", "field_ref": "1", "status_filter": "open"},
            follow_redirects=False)
    page = ui.get(f"/ui/projects/{pid}").text
    # Now it shows in the visible row too (row + modal = twice).
    assert page.count("Must pass SSO test") == 2
    assert "(1 shown)" in page             # filter applied: only the 1 open UC remains

    # Preference persisted per-user.
    pref = db.query(UseCaseViewPref).one()
    cfg = json.loads(pref.config_json)
    assert "success_validation" in cfg["fields"]
    assert cfg["status_filter"] == "open"


def test_dashboard_status_ordering(ui: TestClient) -> None:
    import json

    from app.db import get_session_factory
    from app.models import DashboardPref, ProjectStatus

    db = get_session_factory()()
    statuses = db.query(ProjectStatus).order_by(ProjectStatus.sort_order).all()
    assert len(statuses) >= 2
    reordered = [statuses[1], statuses[0], *statuses[2:]]
    order = ",".join(str(s.id) for s in reordered)

    ui.post("/ui/dashboard/preferences",
            data={"col_name": "1", "sort": "updated", "status_order": order,
                  "status_ids": [str(s.id) for s in statuses]},
            follow_redirects=False)

    pref = db.query(DashboardPref).one()
    cfg = json.loads(pref.config_json)
    assert cfg["status_order"][0] == statuses[1].id

    # Dashboard renders the groups in the saved order.
    page = ui.get("/ui/dashboard").text
    assert page.index(statuses[1].name) < page.index(statuses[0].name)


def test_dark_mode_toggle(ui: TestClient) -> None:
    from app.config import get_settings
    from app.db import get_session_factory
    from app.models import AppUser

    # Default is light, rendered server-side on <html>.
    assert 'data-theme="light"' in ui.get("/ui/dashboard").text

    # Toggle to dark persists to the account and renders on subsequent loads.
    r = ui.post("/ui/theme", data={"theme": "dark"}, follow_redirects=False)
    assert r.status_code == 204
    assert 'data-theme="dark"' in ui.get("/ui/dashboard").text

    username = get_settings().initial_admin_username
    db = get_session_factory()()
    assert db.query(AppUser).filter(AppUser.username == username).one().theme == "dark"

    # Toggling back to light works too.
    ui.post("/ui/theme", data={"theme": "light"}, follow_redirects=False)
    assert 'data-theme="light"' in ui.get("/ui/dashboard").text


def test_user_display_name(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import AppUser, Project

    # Create a user with a display name, then assign them as Sales Engineer.
    ui.post("/ui/settings/admin-users/new",
            data={"username": "rsmith", "password": "password123",
                  "role": "standard", "display_name": "Robby Smith"},
            follow_redirects=False)
    db = get_session_factory()()
    se = db.query(AppUser).filter(AppUser.username == "rsmith").one()
    assert se.display_name == "Robby Smith"

    cid = _create_customer(ui, "Display Co")
    pid = _create_project(ui, cid, "Display POC")
    p = db.get(Project, pid)
    p.sales_engineer_id = se.id
    db.commit()

    # The project page shows the display name, not the username.
    page = ui.get(f"/ui/projects/{pid}").text
    assert "Robby Smith" in page

    # The project edit form's SE dropdown also shows the display name.
    assert "Robby Smith" in ui.get(f"/ui/projects/{pid}/edit").text

    # Editing the display name updates it everywhere (falls back to username when blank).
    ui.post(f"/ui/settings/admin-users/{se.id}/edit",
            data={"display_name": "  "}, follow_redirects=False)
    db.expire_all()
    assert db.get(AppUser, se.id).display_name is None
    assert "rsmith" in ui.get(f"/ui/projects/{pid}").text  # falls back to username


def test_admin_user_role_change(ui: TestClient) -> None:
    from app.config import get_settings
    from app.db import get_session_factory
    from app.models import AppUser

    _make_standard_user("roleuser")
    db = get_session_factory()()
    target = db.query(AppUser).filter(AppUser.username == "roleuser").one()
    assert target.is_admin is False

    # Promote to admin.
    r = ui.post(f"/ui/settings/admin-users/{target.id}/role",
                data={"role": "admin"}, follow_redirects=False)
    assert r.status_code == 303
    db.expire_all()
    assert db.get(AppUser, target.id).is_admin is True

    # Demote back to standard.
    ui.post(f"/ui/settings/admin-users/{target.id}/role",
            data={"role": "standard"}, follow_redirects=False)
    db.expire_all()
    assert db.get(AppUser, target.id).is_admin is False

    # Guard: you can't change your own role (no self lock-out).
    me = db.query(AppUser).filter(
        AppUser.username == get_settings().initial_admin_username).one()
    ui.post(f"/ui/settings/admin-users/{me.id}/role",
            data={"role": "standard"}, follow_redirects=False)
    db.expire_all()
    assert db.get(AppUser, me.id).is_admin is True  # unchanged


# ---------------------------------------------------------------------------
# Lookups + library (admin)
# ---------------------------------------------------------------------------


def test_lookups_index_lists_each_table(ui: TestClient) -> None:
    """The Lookups landing page (reached from Settings) links to all four tables."""
    page = ui.get("/ui/lookups")
    assert page.status_code == 200
    for title in ("Contact Roles", "Project Statuses", "Feature Types", "Use Case Statuses"):
        assert title in page.text
    # And the Settings hub now links to it.
    assert "/ui/lookups" in ui.get("/ui/settings").text


def test_lookup_pages_render(ui: TestClient) -> None:
    for slug in ("contact-roles", "project-statuses", "feature-types", "use-case-statuses"):
        resp = ui.get(f"/ui/lookups/{slug}")
        assert resp.status_code == 200, slug


def test_create_lookup_row(ui: TestClient) -> None:
    resp = ui.post("/ui/lookups/feature-types/new",
                   data={"name": "PAM-Test", "description": "x", "is_active": "1"},
                   follow_redirects=False)
    assert resp.status_code == 303
    assert "PAM-Test" in ui.get("/ui/lookups/feature-types").text


def _default_library_set_id() -> int:
    from app.db import get_session_factory
    from app.models import LibrarySet

    db = get_session_factory()()
    return db.query(LibrarySet).filter(LibrarySet.is_default.is_(True)).one().id


def test_library_admin_create(ui: TestClient) -> None:
    set_id = _default_library_set_id()
    resp = ui.post("/ui/library/new",
                   data={"library_set_id": str(set_id),
                         "category": "Demo", "name": "Library Item X",
                         "default_reference_number": "1.1", "description": "",
                         "success_validation": "", "feature_type_id": ""},
                   follow_redirects=False)
    assert resp.status_code == 303
    assert "Library Item X" in ui.get("/ui/library").text


def _create_library_set(ui: TestClient, name: str, description: str = "") -> int:
    """Create a library via the UI and return its id (parsed from the redirect)."""
    resp = ui.post("/ui/library/sets",
                   data={"name": name, "description": description},
                   follow_redirects=False)
    assert resp.status_code == 303
    return int(resp.headers["location"].split("set=")[1])


def test_library_sets_create_scope_and_isolation(ui: TestClient) -> None:
    standard = _default_library_set_id()
    new_id = _create_library_set(ui, "Acme Launch", "Early adoption")
    assert new_id != standard

    # An entry created in the new library appears only when scoped to it.
    ui.post("/ui/library/new",
            data={"library_set_id": str(new_id), "category": "Onboarding",
                  "name": "Sandbox setup", "default_reference_number": "1.1",
                  "description": "", "success_validation": "", "feature_type_id": ""},
            follow_redirects=False)
    assert "Sandbox setup" in ui.get(f"/ui/library?set={new_id}").text
    assert "Sandbox setup" not in ui.get(f"/ui/library?set={standard}").text

    # The scoped export is named after the library, with date + random token.
    import re as _re
    from datetime import date
    stamp = date.today().strftime("%m%d%Y")
    exp = ui.get(f"/ui/library/export.xlsx?set={new_id}")
    assert exp.status_code == 200
    assert _re.search(
        rf"acme-launch-library-{stamp}-\d{{4}}\.xlsx", exp.headers["content-disposition"]
    )


def test_library_set_delete_blocked_when_non_empty(ui: TestClient) -> None:
    new_id = _create_library_set(ui, "Has Entries")
    ui.post("/ui/library/new",
            data={"library_set_id": str(new_id), "category": "C", "name": "Keeper",
                  "default_reference_number": "", "description": "",
                  "success_validation": "", "feature_type_id": ""},
            follow_redirects=False)
    ui.post(f"/ui/library/sets/{new_id}/delete", follow_redirects=False)
    # Still present — deletion is guarded while the library holds use cases.
    assert "Has Entries" in ui.get("/ui/library/sets").text


def test_library_entry_move_between_sets(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import UseCaseLibrary

    target = _create_library_set(ui, "Destination")
    standard = _default_library_set_id()
    ui.post("/ui/library/new",
            data={"library_set_id": str(standard), "category": "Move", "name": "Mover One",
                  "default_reference_number": "", "description": "",
                  "success_validation": "", "feature_type_id": ""},
            follow_redirects=False)
    db = get_session_factory()()
    entry = db.query(UseCaseLibrary).filter(UseCaseLibrary.name == "Mover One").one()

    # Edit the entry, changing its library — this is the "move" action.
    resp = ui.post(f"/ui/library/{entry.id}/edit",
                   data={"library_set_id": str(target), "category": "Move",
                         "name": "Mover One", "default_reference_number": "",
                         "description": "", "success_validation": "",
                         "feature_type_id": "", "is_active": "1"},
                   follow_redirects=False)
    assert resp.status_code == 303
    db.expire_all()
    assert db.get(UseCaseLibrary, entry.id).library_set_id == target
    assert "Mover One" in ui.get(f"/ui/library?set={target}").text


def _new_library_entry(ui: TestClient, set_id: int, category: str, name: str) -> int:
    from app.db import get_session_factory
    from app.models import UseCaseLibrary

    ui.post("/ui/library/new",
            data={"library_set_id": str(set_id), "category": category, "name": name,
                  "default_reference_number": "", "description": "",
                  "success_validation": "", "feature_type_id": ""},
            follow_redirects=False)
    db = get_session_factory()()
    return db.query(UseCaseLibrary).filter(UseCaseLibrary.name == name).one().id


def test_library_bulk_set_category_and_active(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import UseCaseLibrary

    set_id = _default_library_set_id()
    a = _new_library_entry(ui, set_id, "Old Cat", "Bulk A")
    b = _new_library_entry(ui, set_id, "Old Cat", "Bulk B")

    resp = ui.post("/ui/library/bulk",
                   data={"action": "category", "category": "New Cat",
                         "current_set_id": str(set_id), "ids": [str(a), str(b)]},
                   follow_redirects=False)
    assert resp.status_code == 303
    db = get_session_factory()()
    assert db.get(UseCaseLibrary, a).category == "New Cat"
    assert db.get(UseCaseLibrary, b).category == "New Cat"

    # Bulk deactivate.
    ui.post("/ui/library/bulk",
            data={"action": "active", "is_active": "0",
                  "current_set_id": str(set_id), "ids": [str(a), str(b)]},
            follow_redirects=False)
    db.expire_all()
    assert db.get(UseCaseLibrary, a).is_active is False
    assert db.get(UseCaseLibrary, b).is_active is False


def test_library_bulk_move_and_delete(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import UseCaseLibrary

    src = _default_library_set_id()
    dest = _create_library_set(ui, "Bulk Dest")
    a = _new_library_entry(ui, src, "C", "Move A")
    b = _new_library_entry(ui, src, "C", "Move B")

    # Bulk move to the destination library.
    ui.post("/ui/library/bulk",
            data={"action": "move", "target_set_id": str(dest),
                  "current_set_id": str(src), "ids": [str(a), str(b)]},
            follow_redirects=False)
    db = get_session_factory()()
    assert db.get(UseCaseLibrary, a).library_set_id == dest
    assert db.get(UseCaseLibrary, b).library_set_id == dest

    # Bulk delete them.
    resp = ui.post("/ui/library/bulk",
                   data={"action": "delete", "current_set_id": str(dest),
                         "ids": [str(a), str(b)]},
                   follow_redirects=False)
    assert resp.status_code == 303
    db.expire_all()
    assert db.get(UseCaseLibrary, a) is None
    assert db.get(UseCaseLibrary, b) is None


def test_library_export_and_template_download(ui: TestClient) -> None:
    import re as _re
    from datetime import date

    xlsx_ct = "spreadsheetml"
    stamp = date.today().strftime("%m%d%Y")
    exp = ui.get("/ui/library/export.xlsx")
    assert exp.status_code == 200
    assert xlsx_ct in exp.headers["content-type"]
    # Date + random 4-digit token stamped before the extension.
    assert _re.search(rf"-{stamp}-\d{{4}}\.xlsx", exp.headers["content-disposition"])
    tmpl = ui.get("/ui/library/template.xlsx")
    assert tmpl.status_code == 200
    assert xlsx_ct in tmpl.headers["content-type"]
    # The blank template is the one export that is NOT date/random-stamped.
    assert "use-case-library-template.xlsx" in tmpl.headers["content-disposition"]
    assert stamp not in tmpl.headers["content-disposition"]


def test_library_formatted_xlsx_export(ui: TestClient) -> None:
    import io

    from openpyxl import load_workbook

    import re as _re
    from datetime import date

    r = ui.get("/ui/library/formatted.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    stamp = date.today().strftime("%m%d%Y")
    assert _re.search(rf"use-cases-{stamp}-\d{{4}}\.xlsx", r.headers["content-disposition"])
    # Each export filename is unique (random token) so the browser can't cache it.
    r2 = ui.get("/ui/library/formatted.xlsx")
    assert r.headers["content-disposition"] != r2.headers["content-disposition"]
    ws = load_workbook(io.BytesIO(r.content)).active
    assert ws["A1"].value == "Core Use Case Library"  # title row = library name
    # Header fill must be full-alpha ARGB (FF…) so the brand color actually shows;
    # a bare 6-char hex would get a transparent 00 alpha. Default brand = #1e293b.
    assert ws.cell(row=4, column=1).fill.fgColor.rgb == "FF1E293B"


def test_library_pdf_export(ui: TestClient) -> None:
    # The HTML template must always render (no request, no system libs needed).
    from app.db import get_session_factory
    from app.services import report_pdf
    from app.services.branding import current_branding
    from app.services.library_sets import default_library_set

    db = get_session_factory()()
    lib = default_library_set(db)
    from itertools import groupby

    from app.models import UseCaseLibrary
    entries = (
        db.query(UseCaseLibrary)
        .filter(UseCaseLibrary.library_set_id == lib.id, UseCaseLibrary.is_active.is_(True))
        .order_by(UseCaseLibrary.category, UseCaseLibrary.default_reference_number)
        .all()
    )
    groups = [{"category": c, "entries": list(i)}
              for c, i in groupby(entries, key=lambda e: e.category)]
    html = report_pdf.render_library_html({
        "library": lib, "groups": groups, "total": len(entries), "full": True,
        "branding": current_branding(), "generated_on": "Jan 1, 2026",
    })
    assert lib.name in html
    # Hyperlinked table of contents: each TOC entry links to a use-case anchor.
    assert "Contents" in html
    first = entries[0]
    assert f'href="#uc-{first.id}"' in html
    assert f'id="uc-{first.id}"' in html

    # PDF generation itself needs WeasyPrint's native libs — skip if absent.
    try:
        import weasyprint  # noqa: F401
    except (ImportError, OSError):
        pytest.skip("WeasyPrint or its native libraries are unavailable")
    r = ui.get("/ui/library/export.pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:5] == b"%PDF-"


def test_library_import_roundtrip_updates_and_adds(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.models import UseCaseLibrary

    set_id = _default_library_set_id()

    # Seed one entry to update on re-import.
    ui.post("/ui/library/new",
            data={"library_set_id": str(set_id),
                  "category": "Cat A", "name": "Item One", "default_reference_number": "1.1",
                  "description": "", "success_validation": "", "feature_type_id": ""},
            follow_redirects=False)
    db = get_session_factory()()
    e = db.query(UseCaseLibrary).filter(UseCaseLibrary.name == "Item One").one()

    # CSV matching the template columns: row 0 updates the existing entry (by Id),
    # row 1 has a blank Id so it's added as new.
    csv_text = (
        "Id,Reference Number,Category,Name,Description,Success Validation,Feature Type,Active\n"
        f"{e.id},1.1,Cat A,Item One Renamed,,,,Yes\n"
        ",2.1,Cat B,Brand New Item,A description,,,Yes\n"
    )
    files = {"file": ("lib.csv", csv_text.encode(), "text/csv")}
    prev = ui.post("/ui/library/spreadsheet/preview", files=files,
                   data={"library_set_id": str(set_id)})
    assert prev.status_code == 200
    assert "1 new" in prev.text and "1 update" in prev.text

    # Apply both rows.
    resp = ui.post("/ui/library/spreadsheet/apply", data={
        "library_set_id": str(set_id),
        "select": ["0", "1"],
        "id_0": str(e.id), "category_0": "Cat A", "name_0": "Item One Renamed",
        "ref_0": "1.1", "desc_0": "", "sv_0": "", "feature_type_id_0": "", "active_0": "1",
        "id_1": "", "category_1": "Cat B", "name_1": "Brand New Item",
        "ref_1": "2.1", "desc_1": "A description", "sv_1": "", "feature_type_id_1": "", "active_1": "1",
    }, follow_redirects=False)
    assert resp.status_code == 303

    db.expire_all()
    assert db.get(UseCaseLibrary, e.id).name == "Item One Renamed"  # updated in place
    assert db.query(UseCaseLibrary).filter(UseCaseLibrary.name == "Brand New Item").count() == 1
    page = ui.get("/ui/library").text
    assert "Item One Renamed" in page and "Brand New Item" in page


def test_mcp_token_settings_page_and_rotate(ui: TestClient) -> None:
    from app.services.mcp_token import read_token

    assert ui.get("/ui/settings/mcp").status_code == 200
    assert read_token() is None
    resp = ui.post("/ui/settings/mcp/rotate", follow_redirects=False)
    assert resp.status_code == 303
    # One-time reveal on the next page load, and the token is persisted.
    assert "poct_" in ui.get("/ui/settings/mcp").text
    assert read_token() is not None
    # Clearing revokes + removes it.
    ui.post("/ui/settings/mcp/clear", follow_redirects=False)
    assert read_token() is None


def test_mcp_gateway_token_and_allowed_hosts(ui: TestClient) -> None:
    from app.services import mcp_gateway

    assert ui.get("/ui/settings/mcp").status_code == 200
    assert mcp_gateway.read_gateway_token() is None

    # Generate the inbound gateway token.
    r = ui.post("/ui/settings/mcp/gateway/rotate", follow_redirects=False)
    assert r.status_code == 303
    assert "poctgw_" in ui.get("/ui/settings/mcp").text  # one-time reveal
    assert mcp_gateway.read_gateway_token() is not None

    # Save allowed hosts, then clear them.
    ui.post("/ui/settings/mcp/allowed-hosts",
            data={"allowed_hosts": "mcp.example.com, 10.0.0.5:*"}, follow_redirects=False)
    assert mcp_gateway.read_allowed_hosts() == ["mcp.example.com", "10.0.0.5:*"]
    ui.post("/ui/settings/mcp/allowed-hosts", data={"allowed_hosts": ""}, follow_redirects=False)
    assert mcp_gateway.read_allowed_hosts() == []

    # Clearing the gateway token leaves the endpoint locked down.
    ui.post("/ui/settings/mcp/gateway/clear", follow_redirects=False)
    assert mcp_gateway.read_gateway_token() is None


def test_mcp_key_flagged_and_protected_in_api_keys_list(ui: TestClient) -> None:
    from app.db import get_session_factory
    from app.services import mcp_token

    ui.post("/ui/settings/mcp/rotate", follow_redirects=False)
    db = get_session_factory()()
    mcp_id = mcp_token.current_key_id(db)
    assert mcp_id is not None

    page = ui.get("/ui/settings/api-keys").text
    assert "Manage on MCP page" in page  # dedicated key isn't revoke/delete-able here

    # Revoking/deleting the MCP key from the API-keys list is bounced to the MCP page
    # and leaves the token intact.
    r = ui.post(f"/ui/settings/api-keys/{mcp_id}/revoke", follow_redirects=False)
    assert r.headers["location"] == "/ui/settings/mcp"
    r = ui.post(f"/ui/settings/api-keys/{mcp_id}/delete", follow_redirects=False)
    assert r.headers["location"] == "/ui/settings/mcp"
    assert mcp_token.read_token() is not None


# ---------------------------------------------------------------------------
# RBAC
# ---------------------------------------------------------------------------


def _make_standard_user(username: str = "stduser") -> None:
    from app.db import get_session_factory
    from app.models import AppUser
    from app.services.passwords import hash_password

    db = get_session_factory()()
    db.add(AppUser(username=username, password_hash=hash_password("password123"),
                   is_active=True, is_admin=False))
    db.commit()


def test_standard_user_blocked_from_admin_areas(client: TestClient) -> None:
    _make_standard_user()
    _login(client, "stduser", "password123")
    assert client.get("/ui/projects/", follow_redirects=False).status_code == 200
    for path in ("/ui/library/", "/ui/lookups/contact-roles", "/ui/settings/admin-users"):
        resp = client.get(path, follow_redirects=False)
        assert resp.status_code == 303
        assert resp.headers["location"] == "/ui/dashboard", path


def test_standard_user_can_create_project(client: TestClient) -> None:
    _make_standard_user("stduser2")
    _login(client, "stduser2", "password123")
    cid = _create_customer(client, "StdCo")
    pid = _create_project(client, cid, "StdCo POC")
    assert client.get(f"/ui/projects/{pid}").status_code == 200
