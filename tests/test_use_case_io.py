"""Spreadsheet export / import (deterministic upsert) of use cases."""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.db import get_session_factory
from app.models import (
    Customer,
    FeatureType,
    Project,
    ProjectStatus,
    ProjectUseCase,
    UseCaseStatus,
)

HEADER = [
    "Id", "Reference Number", "Category", "Name", "Description",
    "Success Validation", "Feature Type", "Status", "Comments", "Completed On",
]


def _login(client: TestClient, username: str, password: str) -> None:
    resp = client.post(
        "/ui/login", data={"username": username, "password": password},
        follow_redirects=False,
    )
    assert resp.status_code == 303, resp.text


@pytest.fixture
def admin_ui(client: TestClient) -> TestClient:
    from app.config import get_settings

    s = get_settings()
    _login(client, s.initial_admin_username, s.initial_admin_password)
    return client


_seq = 0


def _project_with_use_cases(names: list[str]) -> tuple[int, list[int]]:
    global _seq
    _seq += 1
    db = get_session_factory()()
    try:
        cust = Customer(name=f"IO Cust {_seq}")
        db.add(cust)
        db.flush()
        ps = db.query(ProjectStatus).order_by(ProjectStatus.sort_order).first()
        us = db.query(UseCaseStatus).order_by(UseCaseStatus.sort_order).first()
        project = Project(customer_id=cust.id, name="IO POC", status_id=ps.id)
        db.add(project)
        db.flush()
        ids = []
        for nm in names:
            uc = ProjectUseCase(
                project_id=project.id, source="custom", category="Access",
                name=nm, status_id=us.id,
            )
            db.add(uc)
            db.flush()
            ids.append(uc.id)
        db.commit()
        return project.id, ids
    finally:
        db.close()


def _xlsx_bytes(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Export & template
# ---------------------------------------------------------------------------


def test_export_roundtrips_ids_and_names(admin_ui: TestClient) -> None:
    pid, ids = _project_with_use_cases(["First UC", "Second UC"])
    resp = admin_ui.get(f"/ui/projects/{pid}/use-cases/export.xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADER
    body = list(ws.iter_rows(min_row=2, values_only=True))
    names = {r[3] for r in body}
    exported_ids = {r[0] for r in body}
    assert names == {"First UC", "Second UC"}
    assert exported_ids == set(ids)  # id column populated for round-trip


def test_template_has_headers_and_dropdowns(admin_ui: TestClient) -> None:
    pid, _ = _project_with_use_cases([])
    resp = admin_ui.get(f"/ui/projects/{pid}/use-cases/template.xlsx")
    assert resp.status_code == 200
    ws = load_workbook(io.BytesIO(resp.content)).active
    assert [c.value for c in ws[1]] == HEADER
    assert len(ws.data_validations.dataValidation) >= 1  # status/feature dropdowns


# ---------------------------------------------------------------------------
# classify_rows (unit)
# ---------------------------------------------------------------------------


def test_classify_new_update_and_warnings(client: TestClient) -> None:
    from app.services.use_case_io import classify_rows

    pid, ids = _project_with_use_cases(["Existing"])
    db = get_session_factory()()
    try:
        project = db.get(Project, pid)
        rows = [
            {"id": str(ids[0]), "reference_number": "1.1", "category": "Access",
             "name": "Existing renamed", "description": "", "success_validation": "",
             "feature_type": "", "status": "Completed", "comments": "", "completed_on": "2026-05-01"},
            {"id": "", "reference_number": "1.2", "category": "Access", "name": "Brand new",
             "description": "", "success_validation": "", "feature_type": "Nope",
             "status": "Bogus", "comments": "", "completed_on": ""},
            {"id": "999999", "reference_number": "", "category": "Access", "name": "Stale id",
             "description": "", "success_validation": "", "feature_type": "", "status": "",
             "comments": "", "completed_on": ""},
            {"id": "", "reference_number": "", "category": "", "name": "",
             "description": "", "success_validation": "", "feature_type": "", "status": "",
             "comments": "", "completed_on": ""},
        ]
        out = classify_rows(db, project, rows)
        assert out[0].action == "update" and out[0].target_id == ids[0]
        assert out[0].completed_on is not None
        assert out[1].action == "new"
        assert any("unknown status" in w for w in out[1].warnings)
        assert any("unknown feature" in w for w in out[1].warnings)
        # stale id (not in this project) → treated as new, flagged
        assert out[2].action == "new" and any("isn't in this project" in w for w in out[2].warnings)
        # blank row → invalid
        assert out[3].valid is False
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Import preview + apply (upsert)
# ---------------------------------------------------------------------------


def test_preview_counts_new_and_update(admin_ui: TestClient) -> None:
    pid, ids = _project_with_use_cases(["Existing"])
    data = _xlsx_bytes([
        [ids[0], "1.1", "Access", "Existing v2", "", "", "", "Completed", "", ""],
        ["", "1.2", "Access", "Fresh one", "", "", "", "", "", ""],
    ])
    resp = admin_ui.post(
        f"/ui/projects/{pid}/use-cases/spreadsheet/preview",
        files={"file": ("import.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    assert "1 new" in resp.text
    assert "1 update" in resp.text


def test_apply_upserts(admin_ui: TestClient) -> None:
    pid, ids = _project_with_use_cases(["Existing"])
    resp = admin_ui.post(
        f"/ui/projects/{pid}/use-cases/spreadsheet/apply",
        data={
            "select": ["0", "1"],
            "id_0": str(ids[0]), "ref_0": "1.1", "category_0": "Access",
            "name_0": "Existing renamed", "desc_0": "", "sv_0": "", "comments_0": "",
            "status_id_0": "", "feature_type_id_0": "", "completed_0": "",
            "id_1": "", "ref_1": "1.2", "category_1": "Access", "name_1": "Brand new",
            "desc_1": "", "sv_1": "", "comments_1": "", "status_id_1": "",
            "feature_type_id_1": "", "completed_1": "",
        },
        follow_redirects=False,
    )
    assert resp.status_code == 303
    db = get_session_factory()()
    try:
        ucs = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).all()
        assert len(ucs) == 2  # one updated in place + one added
        names = {uc.name for uc in ucs}
        assert names == {"Existing renamed", "Brand new"}
        assert db.get(ProjectUseCase, ids[0]).name == "Existing renamed"  # updated, not duplicated
    finally:
        db.close()


def test_apply_with_resolved_status_and_feature(admin_ui: TestClient) -> None:
    pid, _ = _project_with_use_cases([])
    db = get_session_factory()()
    try:
        cid = db.query(UseCaseStatus).filter(UseCaseStatus.is_complete_status.is_(True)).first().id
        ft = db.query(FeatureType).first().id
    finally:
        db.close()
    admin_ui.post(
        f"/ui/projects/{pid}/use-cases/spreadsheet/apply",
        data={
            "select": ["0"], "id_0": "", "ref_0": "9.9", "category_0": "Reporting",
            "name_0": "With status", "desc_0": "", "sv_0": "", "comments_0": "",
            "status_id_0": str(cid), "feature_type_id_0": str(ft), "completed_0": "2026-05-01",
        },
        follow_redirects=False,
    )
    db = get_session_factory()()
    try:
        uc = db.query(ProjectUseCase).filter(ProjectUseCase.project_id == pid).one()
        assert uc.status_id == cid and uc.feature_type_id == ft
        assert uc.completed_on.isoformat() == "2026-05-01"
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Project "tracker" export (report page) — two-tab, formatted, filterable
# ---------------------------------------------------------------------------

TRACKER_HEADERS = [
    "Ref #", "Category", "Name", "Description", "Success validation",
    "Feature", "Status", "Comments", "Completed on",
]


def _project_for_tracker() -> int:
    """A project spanning two categories with one completed use case (dated)."""
    from datetime import date

    global _seq
    _seq += 1
    db = get_session_factory()()
    try:
        cust = Customer(name=f"Trk Cust {_seq}")
        db.add(cust)
        db.flush()
        ps = db.query(ProjectStatus).order_by(ProjectStatus.sort_order).first()
        done = db.query(UseCaseStatus).filter(
            UseCaseStatus.is_complete_status.is_(True)
        ).first()
        open_ = db.query(UseCaseStatus).filter(
            UseCaseStatus.is_complete_status.is_(False)
        ).first()
        project = Project(customer_id=cust.id, name="Tracker POC", status_id=ps.id)
        db.add(project)
        db.flush()
        db.add_all([
            ProjectUseCase(
                project_id=project.id, source="custom", category="Access",
                reference_number="1.1", name="SSO login", status_id=done.id,
                completed_on=date(2026, 6, 14), comments="Verified",
            ),
            ProjectUseCase(
                project_id=project.id, source="custom", category="Reporting",
                reference_number="2.1", name="Export report", status_id=open_.id,
            ),
        ])
        db.commit()
        return project.id
    finally:
        db.close()


def test_tracker_export_has_summary_and_tracker_sheets(admin_ui: TestClient) -> None:
    pid = _project_for_tracker()
    resp = admin_ui.get(f"/ui/reports/projects/{pid}/tracker.xlsx?mode=working")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Summary", "Use case tracker"]

    ws = wb["Use case tracker"]
    assert [c.value for c in ws[1]] == TRACKER_HEADERS
    # Every category appears as a banner row (category text in column A).
    col_a = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "Access" in col_a and "Reporting" in col_a
    # The completed use case keeps a real date in the "Completed on" column.
    from datetime import date, datetime

    dates = [
        c.value for row in ws.iter_rows(min_row=2) for c in (row[8],)
        if isinstance(c.value, (date, datetime))
    ]
    assert date(2026, 6, 14) in [d.date() if isinstance(d, datetime) else d for d in dates]

    ws.auto_filter.ref  # autofilter is set over the grid
    assert ws.auto_filter.ref is not None
    assert ws.freeze_panes == "A2"


def test_tracker_working_is_editable_readonly_is_protected(admin_ui: TestClient) -> None:
    pid = _project_for_tracker()

    working = load_workbook(io.BytesIO(
        admin_ui.get(f"/ui/reports/projects/{pid}/tracker.xlsx?mode=working").content
    ))["Use case tracker"]
    # Working copy: live status dropdown, sheet not protected.
    assert len(working.data_validations.dataValidation) >= 1
    assert working.protection.sheet is False

    readonly = load_workbook(io.BytesIO(
        admin_ui.get(f"/ui/reports/projects/{pid}/tracker.xlsx?mode=readonly").content
    ))["Use case tracker"]
    # Read-only snapshot: protected, no dropdowns.
    assert readonly.protection.sheet is True
    assert len(readonly.data_validations.dataValidation) == 0


def test_csv_import_preview(admin_ui: TestClient) -> None:
    pid, _ = _project_with_use_cases([])
    csv = (
        b"Id,Reference Number,Category,Name,Description,Success Validation,Feature Type,Status,Comments,Completed On\n"
        b",2.1,Reporting,CSV row,desc,,,,,\n"
    )
    resp = admin_ui.post(
        f"/ui/projects/{pid}/use-cases/spreadsheet/preview",
        files={"file": ("import.csv", csv, "text/csv")},
    )
    assert resp.status_code == 200
    assert "CSV row" in resp.text
    assert "1 new" in resp.text
