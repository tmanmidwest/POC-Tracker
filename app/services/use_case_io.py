"""Spreadsheet export/import of a project's use cases (deterministic, no AI).

One column layout is shared by export, the blank template, and import. The
optional ``id`` column drives upsert: a row whose ``id`` matches an existing use
case in the project updates it; a blank ``id`` inserts a new one. Status and
feature type are matched by name (case-insensitive) so the sheet stays readable.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties
from sqlalchemy.orm import Session

from app.models import (
    FeatureType,
    LibrarySet,
    Project,
    ProjectUseCase,
    UseCaseLibrary,
    UseCaseStatus,
)

log = logging.getLogger(__name__)

# Display headers (row 1). Parsing normalizes these (lower + spaces→underscores)
# so "Reference Number" maps to the field key "reference_number".
HEADERS = [
    "Id", "Reference Number", "Category", "Name", "Description",
    "Success Validation", "Feature Type", "Status", "Comments", "Completed On",
]
# Field keys in the same order, used for parsing and round-tripping.
KEYS = [
    "id", "reference_number", "category", "name", "description",
    "success_validation", "feature_type", "status", "comments", "completed_on",
]
MAX_ROWS = 2000


# --- Library layout (master template list) -------------------------------
# The library has no status/comments/completed-on; it does carry an Active flag.
LIBRARY_HEADERS = [
    "Id", "Reference Number", "Category", "Name", "Description",
    "Success Validation", "Feature Type", "Active",
]
LIBRARY_KEYS = [
    "id", "reference_number", "category", "name", "description",
    "success_validation", "feature_type", "active",
]


# ---------------------------------------------------------------------------
# Export & template
# ---------------------------------------------------------------------------


def build_export_xlsx(project: Project) -> bytes:
    """Export the project's current use cases as an .xlsx (with ids, for round-trip)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Use Cases"
    ws.append(HEADERS)
    for uc in _sorted_use_cases(project):
        ws.append(
            [
                uc.id,
                uc.reference_number or "",
                uc.category or "",
                uc.name or "",
                uc.description or "",
                uc.success_validation or "",
                uc.feature_type.name if uc.feature_type else "",
                uc.status.name if uc.status else "",
                uc.comments or "",
                uc.completed_on.isoformat() if uc.completed_on else "",
            ]
        )
    _finish_sheet(ws)
    return _to_bytes(wb)


def build_template_xlsx(db: Session) -> bytes:
    """A blank template: headers, dropdowns for status/feature, and guidance."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Use Cases"
    ws.append(HEADERS)
    ws["A1"].comment = Comment(
        "Leave 'Id' blank for new use cases. On export it's filled in; keep it to "
        "update an existing use case instead of creating a duplicate.",
        "Questlog",
    )
    statuses = [s.name for s in _active(db, UseCaseStatus)]
    features = [f.name for f in _active(db, FeatureType)]
    _add_dropdown(ws, "H", statuses)   # Status column
    _add_dropdown(ws, "G", features)   # Feature Type column
    _finish_sheet(ws)
    return _to_bytes(wb)


def _add_dropdown(ws, col: str, values: list[str]) -> None:
    """Add an Excel list-validation dropdown to a column (rows 2..MAX_ROWS)."""
    joined = ",".join(v for v in values if v and "," not in v)
    if not joined or len(joined) > 250:  # Excel inline-list limit
        return
    dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{MAX_ROWS}")


def _finish_sheet(ws) -> None:
    widths = [6, 16, 20, 30, 48, 36, 16, 18, 28, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sorted_use_cases(project: Project) -> list[ProjectUseCase]:
    return sorted(
        project.use_cases, key=lambda u: (u.category.lower(), u.name.lower())
    )


def _active(db: Session, model) -> list:
    return db.query(model).filter(model.is_active.is_(True)).order_by(model.name).all()


# ---------------------------------------------------------------------------
# Import: parse → classify → apply
# ---------------------------------------------------------------------------


@dataclass
class ImportRow:
    index: int
    action: str  # "new" | "update"
    target_id: int | None
    reference_number: str
    category: str
    name: str
    description: str
    success_validation: str
    comments: str
    status_id: int | None
    status_name: str
    feature_type_id: int | None
    feature_name: str
    completed_on: date | None
    valid: bool
    warnings: list[str] = field(default_factory=list)


class SpreadsheetError(Exception):
    """Raised when a spreadsheet can't be read."""


def parse_spreadsheet(
    filename: str, content: bytes, keys: list[str] | None = None
) -> list[dict]:
    """Parse an .xlsx or .csv into a list of {field_key: str} row dicts.

    ``keys`` is the set of field keys to pull out (defaults to the project
    use-case columns); pass ``LIBRARY_KEYS`` for the library layout.
    """
    keys = keys or KEYS
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content, keys)
    if name.endswith(".csv") or not name:
        return _parse_csv(content, keys)
    # Unknown extension — try xlsx then csv.
    try:
        return _parse_xlsx(content, keys)
    except SpreadsheetError:
        return _parse_csv(content, keys)


def _normalize(header: str) -> str:
    return (header or "").strip().lower().replace(" ", "_")


def _parse_xlsx(content: bytes, keys: list[str]) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetError("Could not read that Excel file.") from exc
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    cols = {_normalize(str(h)): i for i, h in enumerate(header) if h is not None}
    out: list[dict] = []
    for raw in rows_iter:
        row = {k: _cell_str(raw[cols[k]]) if k in cols and cols[k] < len(raw) else "" for k in keys}
        if any(row[k] for k in keys):
            out.append(row)
        if len(out) >= MAX_ROWS:
            break
    wb.close()
    return out


def _parse_csv(content: bytes, keys: list[str]) -> list[dict]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return []
    cols = {_normalize(h): i for i, h in enumerate(header)}
    out: list[dict] = []
    for raw in reader:
        row = {k: (raw[cols[k]].strip() if k in cols and cols[k] < len(raw) else "") for k in keys}
        if any(row[k] for k in keys):
            out.append(row)
        if len(out) >= MAX_ROWS:
            break
    return out


def _cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def classify_rows(db: Session, project: Project, rows: list[dict]) -> list[ImportRow]:
    """Resolve names→ids and decide new-vs-update for each parsed row."""
    statuses = {s.name.lower(): s.id for s in _active(db, UseCaseStatus)}
    features = {f.name.lower(): f.id for f in _active(db, FeatureType)}
    existing_ids = {uc.id for uc in project.use_cases}

    results: list[ImportRow] = []
    for i, row in enumerate(rows):
        warnings: list[str] = []
        name = row["name"].strip()
        category = row["category"].strip()

        target_id = int(row["id"]) if row["id"].strip().isdigit() else None
        if target_id is not None and target_id not in existing_ids:
            warnings.append(f"id {target_id} isn't in this project — will be added as new")
            target_id = None
        action = "update" if target_id is not None else "new"

        status_name = row["status"].strip()
        status_id = statuses.get(status_name.lower()) if status_name else None
        if status_name and status_id is None:
            warnings.append(f"unknown status '{status_name}' — will use the default")

        feature_name = row["feature_type"].strip()
        feature_id = features.get(feature_name.lower()) if feature_name else None
        if feature_name and feature_id is None:
            warnings.append(f"unknown feature type '{feature_name}' — will be left blank")

        completed_on = _parse_date_loose(row["completed_on"])

        valid = bool(name and category)
        if not valid:
            warnings.append("missing name or category — row will be skipped")

        results.append(
            ImportRow(
                index=i, action=action, target_id=target_id,
                reference_number=row["reference_number"].strip(),
                category=category, name=name,
                description=row["description"].strip(),
                success_validation=row["success_validation"].strip(),
                comments=row["comments"].strip(),
                status_id=status_id, status_name=status_name,
                feature_type_id=feature_id, feature_name=feature_name,
                completed_on=completed_on, valid=valid, warnings=warnings,
            )
        )
    return results


def _parse_date_loose(value: str) -> date | None:
    value = (value or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


# ===========================================================================
# Library import/export (master template list — no project, no status)
# ===========================================================================

_TRUE_WORDS = {"yes", "y", "true", "1", "active"}
_FALSE_WORDS = {"no", "n", "false", "0", "inactive"}


def _bool_str(value: bool) -> str:
    return "Yes" if value else "No"


def _parse_bool(value: str, default: bool = True) -> bool:
    v = (value or "").strip().lower()
    if v in _TRUE_WORDS:
        return True
    if v in _FALSE_WORDS:
        return False
    return default


def build_library_export_xlsx(db: Session, library_set_id: int | None = None) -> bytes:
    """Export a library as an .xlsx (with ids, for round-trip update).

    Scoped to ``library_set_id`` when given; otherwise exports every library.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Library"
    ws.append(LIBRARY_HEADERS)
    query = db.query(UseCaseLibrary)
    if library_set_id is not None:
        query = query.filter(UseCaseLibrary.library_set_id == library_set_id)
    entries = query.order_by(
        UseCaseLibrary.category,
        UseCaseLibrary.default_reference_number,
        UseCaseLibrary.name,
    ).all()
    for e in entries:
        ws.append(
            [
                e.id,
                e.default_reference_number or "",
                e.category or "",
                e.name or "",
                e.description or "",
                e.success_validation or "",
                e.feature_type.name if e.feature_type else "",
                _bool_str(e.is_active),
            ]
        )
    _finish_library_sheet(ws)
    return _to_bytes(wb)


def build_library_template_xlsx(db: Session) -> bytes:
    """A blank library template: headers, a Feature Type dropdown, and guidance."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Library"
    ws.append(LIBRARY_HEADERS)
    ws["A1"].comment = Comment(
        "Leave 'Id' blank for new entries. On export it's filled in; keep it to "
        "update an existing library entry instead of creating a duplicate.",
        "Questlog",
    )
    _add_dropdown(ws, "G", [f.name for f in _active(db, FeatureType)])  # Feature Type
    _add_dropdown(ws, "H", ["Yes", "No"])  # Active
    _finish_library_sheet(ws)
    return _to_bytes(wb)


def _finish_library_sheet(ws) -> None:
    widths = [6, 16, 20, 30, 48, 36, 16, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


@dataclass
class LibraryImportRow:
    index: int
    action: str  # "new" | "update"
    target_id: int | None
    reference_number: str
    category: str
    name: str
    description: str
    success_validation: str
    feature_type_id: int | None
    feature_name: str
    is_active: bool
    valid: bool
    warnings: list[str] = field(default_factory=list)


def classify_library_rows(
    db: Session, rows: list[dict], library_set_id: int | None = None
) -> list[LibraryImportRow]:
    """Resolve feature-type names→ids and decide new-vs-update for each row.

    When ``library_set_id`` is given, only ids already in that library count as
    updates; an id from another library is treated as a new entry (and added to
    the target library), so an import never silently moves entries between
    libraries.
    """
    features = {f.name.lower(): f.id for f in _active(db, FeatureType)}
    id_query = db.query(UseCaseLibrary.id)
    if library_set_id is not None:
        id_query = id_query.filter(UseCaseLibrary.library_set_id == library_set_id)
    existing_ids = {eid for (eid,) in id_query.all()}

    results: list[LibraryImportRow] = []
    for i, row in enumerate(rows):
        warnings: list[str] = []
        name = row["name"].strip()
        category = row["category"].strip()

        target_id = int(row["id"]) if row["id"].strip().isdigit() else None
        if target_id is not None and target_id not in existing_ids:
            warnings.append(f"id {target_id} isn't in this library — will be added as new")
            target_id = None
        action = "update" if target_id is not None else "new"

        feature_name = row["feature_type"].strip()
        feature_id = features.get(feature_name.lower()) if feature_name else None
        if feature_name and feature_id is None:
            warnings.append(f"unknown feature type '{feature_name}' — will be left blank")

        valid = bool(name and category)
        if not valid:
            warnings.append("missing name or category — row will be skipped")

        results.append(
            LibraryImportRow(
                index=i, action=action, target_id=target_id,
                reference_number=row["reference_number"].strip(),
                category=category, name=name,
                description=row["description"].strip(),
                success_validation=row["success_validation"].strip(),
                feature_type_id=feature_id, feature_name=feature_name,
                is_active=_parse_bool(row["active"], default=True),
                valid=valid, warnings=warnings,
            )
        )
    return results


# ===========================================================================
# Library "presentation" export — a styled, read-only .xlsx for sharing
# (distinct from the plain round-trip export above, which import re-reads).
# ===========================================================================

_PRES_HEADERS = ["Ref", "Name", "Description", "Success Validation", "Feature Type"]
_PRES_WIDTHS = [12, 34, 55, 45, 18]


def _accent_hex() -> str:
    """The brand accent as 8-char ARGB hex (for openpyxl fills/fonts).

    openpyxl needs a full ARGB value; a bare 6-char RGB gets a ``00`` (fully
    transparent) alpha prepended, so the fill would silently render invisible.
    """
    from app.services.branding import current_branding

    raw = (current_branding().get("color") or "#1E293B").lstrip("#")
    rgb = raw.upper() if len(raw) == 6 else "1E293B"
    return "FF" + rgb


def build_library_presentation_xlsx(
    db: Session, library_set: LibrarySet, *, active_only: bool = True
) -> bytes:
    """A polished, read-only .xlsx of one library: title, branded headers,
    category banners, wrapped text, and borders. Grouped by category."""
    accent = _accent_hex()  # 8-char ARGB
    header_fill = PatternFill("solid", fgColor=accent)
    banner_fill = PatternFill("solid", fgColor="FFF1F5F9")
    white_bold = Font(bold=True, color="FFFFFFFF")
    muted = Font(color="FF64748B")
    title_font = Font(bold=True, size=16, color=accent)
    banner_font = Font(bold=True, color="FF0F172A")
    thin = Side(style="thin", color="FFE2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")

    query = db.query(UseCaseLibrary).filter(
        UseCaseLibrary.library_set_id == library_set.id
    )
    if active_only:
        query = query.filter(UseCaseLibrary.is_active.is_(True))
    entries = query.order_by(
        UseCaseLibrary.category,
        UseCaseLibrary.default_reference_number,
        UseCaseLibrary.name,
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Use Cases"
    ncols = len(_PRES_HEADERS)
    last_col = chr(ord("A") + ncols - 1)

    # Title + subtitle (merged across all columns).
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = library_set.name
    ws["A1"].font = title_font
    ws.merge_cells(f"A2:{last_col}2")
    sub = f"Use Case Library · {len(entries)} use case{'' if len(entries) == 1 else 's'}"
    if library_set.description:
        sub += f" · {library_set.description}"
    ws["A2"] = sub
    ws["A2"].font = muted

    # Header row (row 4).
    header_row = 4
    for i, head in enumerate(_PRES_HEADERS):
        cell = ws.cell(row=header_row, column=i + 1, value=head)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(vertical="center")
        cell.border = border

    # Body: a category banner, then its entries.
    r = header_row + 1
    last_category = object()
    for e in entries:
        if e.category != last_category:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
            banner = ws.cell(row=r, column=1, value=e.category)
            banner.fill = banner_fill
            banner.font = banner_font
            banner.alignment = Alignment(vertical="center")
            r += 1
            last_category = e.category
        values = [
            e.default_reference_number or "",
            e.name or "",
            e.description or "",
            e.success_validation or "",
            e.feature_type.name if e.feature_type else "",
        ]
        for i, val in enumerate(values):
            cell = ws.cell(row=r, column=i + 1, value=val)
            cell.border = border
            cell.alignment = wrap_top if i in (2, 3) else top
        r += 1

    for i, w in enumerate(_PRES_WIDTHS):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    ws.freeze_panes = f"A{header_row + 1}"
    return _to_bytes(wb)


# ===========================================================================
# Project "tracker" export — a polished, two-tab .xlsx for the report page.
# A dashboard "Summary" sheet plus a color-coded, filterable "Use case
# tracker" sheet grouped by category. Two flavors: a working copy (status
# dropdowns, unlocked, editable) and a read-only snapshot (protected).
# ===========================================================================

_TRACKER_HEADERS = [
    "Ref #", "Category", "Name", "Description", "Success validation",
    "Feature", "Status", "Comments", "Completed on",
]
_TRACKER_WIDTHS = [8, 18, 30, 50, 40, 14, 16, 34, 14]
_TRACKER_STATUS_COL = 7  # 1-based index of the "Status" column (G)

# Status → (fill, font) by bucket. Chosen to echo the on-screen report badges.
_STATUS_STYLES = {
    "done": ("FFDCFCE7", "FF15803D"),
    "progress": ("FFFEF3C7", "FF92400E"),
    "blocked": ("FFFEE2E2", "FF991B1B"),
    "pending": ("FFF1F5F9", "FF475569"),
}


def _status_bucket(status) -> str:
    """Classify a use-case status for coloring and the summary counts.

    ``is_complete_status`` is authoritative for "done"; the rest are inferred
    from the (user-configurable) status name, defaulting to "pending".
    """
    if status is None:
        return "pending"
    if getattr(status, "is_complete_status", False):
        return "done"
    name = (status.name or "").lower()
    if any(w in name for w in ("block", "hold", "fail", "reject")):
        return "blocked"
    if any(w in name for w in ("progress", "wip", "started", "testing", "active")):
        return "progress"
    return "pending"


def build_project_tracker_xlsx(
    db: Session, project: Project, groups: list[dict], *, editable: bool = True
) -> bytes:
    """A polished two-tab tracker .xlsx for a project's use cases.

    ``groups`` is the report's category-grouped use cases (list of
    ``{"category", "use_cases", ...}`` dicts). ``editable`` toggles a working
    copy (status dropdowns, unlocked cells) versus a read-only snapshot
    (protected sheets, no dropdowns) — both share the same look.
    """
    accent = _accent_hex()  # 8-char ARGB
    header_fill = PatternFill("solid", fgColor=accent)
    banner_fill = PatternFill("solid", fgColor="FFF1F5F9")
    white_bold = Font(bold=True, color="FFFFFFFF")
    banner_font = Font(bold=True, color="FF0F172A")
    muted = Font(color="FF64748B")
    thin = Side(style="thin", color="FFE2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    top = Alignment(vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    _build_tracker_summary(wb.active, project, groups, accent, header_fill, white_bold, muted)
    _build_tracker_sheet(
        wb.create_sheet("Use case tracker"), db, groups, editable,
        header_fill, white_bold, banner_fill, banner_font, border, wrap_top, top, center,
    )
    return _to_bytes(wb)


def _build_tracker_summary(
    ws, project, groups, accent, header_fill, white_bold, muted
) -> None:
    ws.title = "Summary"
    all_ucs = [uc for g in groups for uc in g["use_cases"]]
    counts = {"done": 0, "progress": 0, "blocked": 0, "pending": 0}
    for uc in all_ucs:
        counts[_status_bucket(uc.status)] += 1
    total = len(all_ucs)
    pct = counts["done"] / total if total else 0

    # Branded title band (rows 1–2).
    ws.merge_cells("A1:E1")
    ws["A1"] = project.display_name
    ws["A1"].fill = header_fill
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFFFF")
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:E2")
    ws["A2"] = f"Use case tracker · generated {date.today().strftime('%b %-d, %Y')}"
    ws["A2"].fill = header_fill
    ws["A2"].font = Font(color="FFDBEAFE")
    ws["A2"].alignment = Alignment(vertical="center", indent=1)

    # Metric row (labels row 4, values row 5).
    metrics = [
        ("Total", total, "FF0F172A"),
        ("Completed", counts["done"], _STATUS_STYLES["done"][1]),
        ("In progress", counts["progress"], _STATUS_STYLES["progress"][1]),
        ("Pending / blocked", counts["pending"] + counts["blocked"], _STATUS_STYLES["blocked"][1]),
        ("Complete", pct, accent),
    ]
    for i, (label, value, color) in enumerate(metrics, start=1):
        lbl = ws.cell(row=4, column=i, value=label)
        lbl.font = muted
        val = ws.cell(row=5, column=i, value=value)
        val.font = Font(bold=True, size=18, color=color)
        if label == "Complete":
            val.number_format = "0%"

    # Per-category breakdown table (starts row 7).
    head_row = 7
    for i, head in enumerate(["Category", "Done", "Total", "Complete"], start=1):
        cell = ws.cell(row=head_row, column=i, value=head)
        cell.fill = header_fill
        cell.font = white_bold
    r = head_row + 1
    for g in groups:
        ucs = g["use_cases"]
        done = sum(1 for uc in ucs if _status_bucket(uc.status) == "done")
        ws.cell(row=r, column=1, value=g["category"])
        ws.cell(row=r, column=2, value=done)
        ws.cell(row=r, column=3, value=len(ucs))
        pcell = ws.cell(row=r, column=4, value=(done / len(ucs) if ucs else 0))
        pcell.number_format = "0%"
        r += 1

    for i, w in enumerate([28, 10, 10, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def _build_tracker_sheet(
    ws, db, groups, editable,
    header_fill, white_bold, banner_fill, banner_font, border, wrap_top, top, center,
) -> None:
    ncols = len(_TRACKER_HEADERS)

    # Header row (row 1), with autofilter and frozen pane below it.
    for i, head in enumerate(_TRACKER_HEADERS, start=1):
        cell = ws.cell(row=1, column=i, value=head)
        cell.fill = header_fill
        cell.font = white_bold
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 22

    # Body: an unmerged colored banner per category, then its use cases as
    # collapsible (outline level 1) rows. Banners stay unmerged so autofilter
    # and sorting keep working; the label just overflows the empty cells.
    r = 2
    for g in groups:
        for c in range(1, ncols + 1):
            bcell = ws.cell(row=r, column=c)
            bcell.fill = banner_fill
            bcell.border = border
        label = ws.cell(row=r, column=1, value=g["category"])
        label.font = banner_font
        label.alignment = Alignment(vertical="center", indent=1)
        r += 1
        for uc in g["use_cases"]:
            _write_tracker_row(ws, r, uc, border, wrap_top, top, center)
            ws.row_dimensions[r].outlineLevel = 1
            r += 1
    last_row = r - 1

    ws.sheet_properties.outlinePr.summaryBelow = False  # banners sit above their rows
    if last_row >= 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_row}"

    for i, w in enumerate(_TRACKER_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Landscape, fit-to-width, repeat the header row on every printed page.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"

    if editable and last_row >= 2:
        # Live status dropdown on the working copy.
        statuses = [s.name for s in _active(db, UseCaseStatus)]
        col = get_column_letter(_TRACKER_STATUS_COL)
        _add_dropdown_range(ws, f"{col}2:{col}{last_row}", statuses)
    elif not editable:
        # Read-only snapshot: lock the sheet but keep filtering usable.
        ws.protection.sheet = True
        ws.protection.autoFilter = False
        ws.protection.selectLockedCells = False


def _write_tracker_row(ws, r, uc, border, wrap_top, top, center) -> None:
    values = [
        uc.reference_number or "",
        uc.category or "",
        uc.name or "",
        uc.description or "",
        uc.success_validation or "",
        uc.feature_type.name if uc.feature_type else "",
        uc.status.name if uc.status else "",
        uc.comments or "",
    ]
    for i, val in enumerate(values, start=1):
        cell = ws.cell(row=r, column=i, value=val)
        cell.border = border
        cell.alignment = wrap_top if i in (4, 5, 8) else top
    # Completed on as a real date so Excel sorts/formats it.
    dcell = ws.cell(row=r, column=9, value=uc.completed_on)
    dcell.border = border
    dcell.alignment = top
    if uc.completed_on:
        dcell.number_format = "yyyy-mm-dd"
    # Color-code the status cell by bucket.
    fill_hex, font_hex = _STATUS_STYLES[_status_bucket(uc.status)]
    scell = ws.cell(row=r, column=_TRACKER_STATUS_COL)
    scell.fill = PatternFill("solid", fgColor=fill_hex)
    scell.font = Font(bold=True, color=font_hex)
    scell.alignment = center


def _add_dropdown_range(ws, cell_range: str, values: list[str]) -> None:
    """Add an Excel list-validation dropdown to an explicit range."""
    joined = ",".join(v for v in values if v and "," not in v)
    if not joined or len(joined) > 250:  # Excel inline-list limit
        return
    dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)
